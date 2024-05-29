'''
 fmt = getattr(settings, 'LOG_FORMAT', None)
 lvl = getattr(settings, 'LOG_LEVEL', logging.DEBUG)

 logging.basicConfig(format=fmt, level=lvl)
 logging.debug(nemailbject.OrderId.id)
 '''
from django.shortcuts import render
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.db.models import Sum
import jdatetime
from django.http import HttpResponseRedirect
from django.http import HttpResponse
from django.views.decorators import csrf
import django.core.serializers
import logging
from django.conf import settings

from myapp.models.fish import *
from myapp.models.message import *
#from django.core import serializers
import json
from django.forms.models import model_to_dict
from myapp.forms import MessageForm
from django.urls import reverse_lazy,reverse
from django.db import transaction
from django.contrib.auth.decorators import login_required
from django.core.paginator import *
from openpyxl import load_workbook
from django.db.models import Q
import os
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import permission_required
from PyPDF2 import PdfReader, PdfWriter
from django.core.files.base import ContentFile
def doPaging(request,books):
    page=request.GET.get('page',1)
    paginator = Paginator(books, 8)
    wos=None
    try:
        wos=paginator.page(page)
    except PageNotAnInteger:
        wos = paginator.page(1)
    except EmptyPage:
        wos = paginator.page(paginator.num_pages)
    return wos

@login_required
@permission_required('myapp.view_fish')
def list_fish(request):
    #
    books=Fish.objects.all()
    wos=doPaging(request,(books))


    return render(request, 'myapp/fish/fishList.html', {'fishes': wos,'title':'فیش حقوقی','section':'list_mail'})
def search_fish(request):
    q=request.GET.get("q",False)
    data=dict()
    if(q):
    #
        books=Fish.objects.filter(Q(name__contains=q)|Q(code=q)|Q(code_meli=q))
        wos=doPaging(request,(books))
        data['result']= render_to_string('myapp/fish/partialFishList.html', {
                'fishes': wos
            })
        return JsonResponse(data)

    books=Fish.objects.all()
    wos=doPaging(request,(books))
    data['result']= render_to_string('myapp/fish/partialFishList.html', {
            'fishes': wos
        })
    return JsonResponse(data)
def details_fish(request):
    code_meli=request.POST.get('code_meli',False)
    code=request.POST.get('code',False)
    mah=request.POST.get('mah1',1)
    sal=request.POST.get('year',1403)
    if(int(sal)==1403):
        try:

            person=Person.objects.get(pid=code,codemeli=code_meli)




            month=['','فروردین','ارديبهشت','خرداد','تیر','مرداد','شهریور','مهر','آبان','آذر','دی','بهمن','اسفند']
            sal2=Salary.objects.get(mah=mah,sal=1403,person=person)


            context = {
                    # 'zipped_lists': zipped_lists,
                    'person':person,
                    # 'last1':last1,
                    # 'last2':last2,
                    # 'last3':last3,
                    'sal':sal,
                    'mah':month[int(mah)],
                    'pdf':sal2.fishFile
                }
            return render(request, 'myapp/fish/fishdetaile2.html', context)
        except Exception as e:
                print(e)
                return render(request, 'myapp/page-error-404.html', {'ss':''})
    else:

        print("mah1",mah1)
        month=['','فروردین','ارديبهشت','خرداد','تیر','مرداد','شهریور','مهر','آبان','آذر','دی','بهمن','اسفند']

        wos=Fish.objects.filter(code=code,code_meli=code_meli,mah=mah,sal=sal)
        if(wos.count()>0):
            karkard=int(wos[0].monthly_hoghugh)/int(wos[0].daily_hoghugh)

            return render(request, 'myapp/fish/fishDetal.html', {'c': wos[0],'karkard':karkard,'title':'صندوق دریافت','section':'list_mail','mah':month[int(wos[0].mah)],'sal':wos[0].sal})
        return render(request, 'myapp/page-error-404.html', {'ss':''})
@login_required
@transaction.atomic
def fish_upload(request):
    def iter_rows(ws):
        for row in ws.iter_rows():
            yield [cell.value for cell in row]
    if request.method == 'POST':
        # print("here!!!",request.FILES)
        my_file=request.FILES.get('file')
        # print(my_file.name)
        month=request.POST.get('month_select',False)
        sal=request.POST.get('sal_select',False)




        msg=MessageFile.objects.create(msgFile=my_file)
        workbook = load_workbook(filename='media/'+msg.msgFile.name)
        ws = workbook.active
        # print(list(iter_rows(ws))[1])
        # Fish.objects.all().delete()
        item_new=Fish(pk=None)
        item_old=Fish(pk=None)

        for i in list(iter_rows(ws)):
            if(i[20]==None):
                pass
            elif(not str(i[30]).isdigit()):
                pass
            else:
                if(i[30] != None):
                    item_new.pk=None
                    item_new.mah=month
                    item_new.sal=sal


                    item_new.mande_morakhasi="{0}:{1}:{2}".format(i[32],i[33],i[34])
                    item_new.code=i[30] if(i[30]) else 0
                    item_new.code_meli=i[29] if(i[29]) else 0
                    item_new.name=i[28] if(i[28]) else 0
                    item_new.daily_hoghugh=i[27] if(i[27]) else 0
                    item_new.karkard=i[26] if(i[26]) else 0
                    item_new.monthly_hoghugh=i[25] if(i[25]) else 0
                    item_new.bon=i[24] if(i[24]) else 0
                    item_new.maskan=i[23] if(i[23]) else 0
                    item_new.nobate_kari=i[22] if(i[22]) else 0
                    item_new.paye_sanavat=i[21] if(i[21]) else 0
                    item_new.ovlad_moavaghe=i[20] if(i[20]) else 0
                    item_new.ovlad=i[19] if(i[19]) else 0
                    item_new.ezafe_kar=i[18] if(i[18]) else 0
                    item_new.padash=i[17] if(i[17]) else 0
                    item_new.eslahe_hoghugh=i[16] if(i[16]) else 0
                    item_new.randeman=i[15] if(i[15]) else 0
                    item_new.jome_kari=i[14] if(i[14]) else 0
                    item_new.ravande_mahe_ghabl=i[13] if(i[13]) else 0
                    item_new.maliat=i[12] if(i[12]) else 0
                    item_new.haghe_bime=i[11] if(i[11]) else 0
                    item_new.mosaede=i[10] if(i[10]) else 0
                    item_new.kasre_hoghugh=i[9] if(i[9]) else 0
                    item_new.vame_dakheli=i[8] if(i[8]) else 0
                    item_new.bime_takmili=i[7] if(i[7]) else 0
                    item_new.vame_aghsati=i[6] if(i[6]) else 0
                    item_new.bime_azad=i[5] if(i[5]) else 0
                    item_new.jarime=i[4] if(i[4]) else 0
                    item_new.pool_khurd=i[3] if(i[3]) else 0
                    item_new.morakhasi=i[2] if(i[2]) else 0
                    item_new.ezafe_kar_time=i[1] if(i[1]) else 0
                    item_new.estelaji=i[0] if(i[0]) else 0
                    item_new.haghe_tahol=i[36] if(i[36]) else 0
                    item_new.tatil_kar=i[35] if(i[35]) else 0
                    item_new.save()




        data=dict()
        data["id"]=msg.id
        return JsonResponse(data)
        # return HttpResponse('')
    return JsonResponse({'post':'fasle'})
@permission_required('myapp.view_fish')
def fish_create(request):
    if (request.method == 'POST'):

        return HttpResponseRedirect(reverse('list_mail'))
    else:

        # form = MessageForm({'fromUser':SysUser.objects.get(userId=request.user.id),'messageStatus':2})
        # form.isupdated=False
        # return save_mail_form(request, form, 'myapp/mail/partialMailCreate.html')
        return render(request, 'myapp/fish/partialFishCreate.html', {'form':''})
def fish_view(request,id):
    pass
def register_shekayat(reuqest,id):
    fish=Fish.objects.get(id=id)
    shekayat=request.POST.get('shekayat',False)
    if(shekayat):
        fish.shekayat=shekayat
        fish.save()

        return HttpResponseRedirect(reverse('index'))
def find_next_person_index(current_row,ws):
            next_row=current_row
            for row in range(current_row + 1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=13).value
                if cell_value and isinstance(cell_value, str):  # Check if the cell has a non-empty string value
                    next_row = row
                    break
            return next_row
def find_person_by_id(pid_str,name,codemeli):
    if(len(codemeli)==0):
        codemeli='1111111111'

    person_exists = Person.objects.filter(pid=pid_str).exists()
    if person_exists:
        return Person.objects.get(pid=pid_str)
    else:

        return Person.objects.create(pid=pid_str,name=name,codemeli=codemeli)
def get_file_extension(file):
            file_name = file.name
            return file_name.split('.')[-1] if '.' in file_name else None
def processfile(request):
    month=request.GET.get('month',False)
    mah=month
    sal=request.GET.get('sal',False)

    msg_pdf=MessageFile2.objects.filter(sal=sal,mah=mah)[0]
    msg_excel=MessageFile.objects.filter(sal=sal,mah=mah)[0]

    workbook = load_workbook(filename='media/'+msg_excel.msgFile.name)
    ws = workbook.active
    m_col=ws.max_column
    m_row = ws.max_row
    index=0
    for i in range(1, m_row + 1):
            # print(ws[i][0].value)
            try:
                value = ws[i][13].value
                if(value):
                        print(value,',',ws[i+5][25].value)
                        pid=find_person_by_id(ws[i+5][25].value,ws[i+8][25].value +' '+ ws[i+11][25].value,'11111')
                        next_person_index=find_next_person_index(i,ws)

                        if(i<next_person_index):
                           p=Salary.objects.get(index=index,mah=mah,sal=sal)
                           p.person=pid
                           p.save()

                        elif(i==next_person_index):
                            p=Salary.objects.get(index=index,mah=mah,sal=sal)
                            p.person=pid
                            p.save()
                        index+=1

            except Exception as ex:
                 print(ex)
                 print(index)



    data=dict()
    data["form_is_valid"]=True
    return JsonResponse(data)



@login_required
def fish_upload2(request):


    if request.method == 'POST':

        my_file=request.FILES.get('file')

        month=request.POST.get('month_select',False)

        mah=month
        sal=request.POST.get('sal_select',False)
        if(get_file_extension(my_file)=="pdf"):
            MessageFile2.objects.filter(mah=mah,sal=sal).delete()
            msg=MessageFile2.objects.create(msgFile=my_file,mah=mah,sal=sal)
            with open('media/'+msg.msgFile.name, 'rb') as pdf_file:
                pdf_reader = PdfReader(pdf_file)

                # Iterate through each page and save as a separate PDF
                for page_num in range(len(pdf_reader.pages)):
                    pdf_writer = PdfWriter()
                    pdf_writer.add_page(pdf_reader.pages[page_num])

                    # Save each page as a separate PDF file
                    output_filename = f'page_{page_num + 1}.pdf'  # Naming each page file
                    with open(output_filename, 'wb') as output_file:
                        pdf_writer.write(output_file)
                        # Salary.objects.create(index=page_num,fishFile=output_file,mah=mah,sal=sal)
                    with open(output_filename, 'rb') as output_file:
                        # Read the content of the file
                        file_content = output_file.read()
                        salary_instance = Salary.objects.create(index=page_num, mah=mah, sal=sal)
                        salary_instance.fishFile.save('desired_filename.pdf', ContentFile(file_content), save=True)
                        # print(f"Page {page_num + 1} saved as '{output_filename}'")
        elif(get_file_extension(my_file)=="xlsx"):
            # print("Excel file")
            MessageFile.objects.filter(mah=mah,sal=sal).delete()
            msg=MessageFile.objects.create(msgFile=my_file,mah=mah,sal=sal)
            # with open('media/'+msg.msgFile.name, 'rb') as pdf_file:
            #     pdf_reader = PdfReader(pdf_file)

            #     # Iterate through each page and save as a separate PDF
            #     for page_num in range(len(pdf_reader.pages)):
            #         pdf_writer = PdfWriter()
            #         pdf_writer.add_page(pdf_reader.pages[page_num])

            #         # Save each page as a separate PDF file
            #         output_filename = f'page_{page_num + 1}.pdf'  # Naming each page file
            #         with open(output_filename, 'wb') as output_file:
            #             pdf_writer.write(output_file)
            #         print(f"Page {page_num + 1} saved as '{output_filename}'")
        # workbook = load_workbook(filename='media/'+msg.msgFile.name)
        # ws = workbook.active
        # m_col=ws.max_column
        # m_row = ws.max_row
        # for i in range(1, m_row + 1):
        #     try:
        #         value = ws[i][10].value
        #         if(value):
        #             # pid=find_person_by_id(ws[i][0].value,value)
        #             next_person_index=find_next_person_index(i,ws)

        #             if(i<next_person_index):
        #                 upload_file_async.delay(i,next_person_index,mah,sal,msg.msgFile.name,ws[i][0].value,value)
        #             elif(i==next_person_index):
        #                 upload_file_async.delay(i,m_row,mah,sal,msg.msgFile.name,ws[i][0].value,value)



        #     except Exception as e:
        #         print(e)  # it was a string, not an int.
        #     workbook.close()



        data=dict()
        data["id"]=msg.id
        return JsonResponse(data)
        # return HttpResponse('')
    return JsonResponse({'post':'fasle'})
@login_required
@csrf_exempt
def fish_delete(request,id):
    comp1 = get_object_or_404(MessageFile, id=id)
    page=request.GET.get('page',1)
    data = dict()
    if (request.method == 'POST'):
        Salary.objects.filter(sal=comp1.sal,mah=comp1.mah).delete()
        MessageFile2.objects.filter(mah=comp1.mah,sal=comp1.sal).delete()
        comp1.delete()
        data['form_is_valid'] = True  # This is just to play along with the existing code
        books=MessageFile.objects.all().order_by('-sal','-mah')
        wos=doPaging(request,books)

        #Tasks.objects.filter(woId=id).update(workorder=id)
        data['html_wo_list'] = render_to_string('myapp/fish/partialFishList.html', {
            'fishes': wos,
            'perms': PermWrapper(request.user),
            'page':page
        })


    return JsonResponse(data)
